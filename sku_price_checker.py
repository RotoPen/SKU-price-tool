import streamlit as st
import pandas as pd
import numpy as np
import os
import json
import openpyxl
from io import BytesIO
import io
from st_aggrid import AgGrid, GridOptionsBuilder, JsCode

pd.options.display.float_format = '{:,.0f}'.format

# 字段名映射（请根据实际表头调整）
SKU_FIELD = "SKU"  # SKU表中的SKU字段
PARENT_SKU_FIELD = "Parent SKU"  # SKU表中的Parent SKU字段
TOOL_SKU_FIELD = "sku编码"  # 工具价格表中的sku编码字段
TOOL_PRICE_FIELD = "活动价格"  # 工具价格表中的活动价格字段
CAMPAIGN_PRODUCT_ID = "Product ID"
CAMPAIGN_VARIATION_ID = "Variation ID"
CAMPAIGN_PRICE_FIELD = "Campaign Price"
CAMPAIGN_RECOMMEND_FIELD = "Recommended Campaign Price"

def strip_columns(df):
    if df is not None:
        df.columns = [str(col).strip() for col in df.columns]
    return df

def clean_id_column(df, col):
    if df is not None and col in df.columns:
        df[col] = df[col].astype(str).str.strip().str.replace('.0', '', regex=False)
    return df

# 新增：同步价格数据的辅助函数，避免重复代码
def sync_price_data(campaign_df, price_input_df, key_columns, value_columns=None, update_price_source=False):
    """
    将price_input_df中的数据同步到campaign_df中
    
    参数:
    campaign_df: 目标DataFrame
    price_input_df: 源DataFrame
    key_columns: 用于匹配两个DataFrame的键列
    value_columns: 需要同步的值列，默认为None时将自动确定
    update_price_source: 是否更新价格来源，默认为False
    
    返回:
    更新后的campaign_df
    """
    if value_columns is None:
        value_columns = [CAMPAIGN_PRICE_FIELD]
        # 检查其它可能的值列是否存在
        for col in ['已修改', '价格有效', '已人工确认']:
            if col in price_input_df.columns:
                value_columns.append(col)
    
    # 确保所有值列都存在于price_input_df中
    existing_columns = [col for col in value_columns if col in price_input_df.columns]
    if len(existing_columns) < len(value_columns):
        missing = set(value_columns) - set(existing_columns)
        st.warning(f"同步数据时发现缺失列: {', '.join(missing)}")
        value_columns = existing_columns
    
    # 如果'已人工确认'不在campaign_df但需要同步，添加该列
    if '已人工确认' in value_columns and '已人工确认' not in campaign_df.columns:
        campaign_df['已人工确认'] = False
        
    # 如果'已修改'不在campaign_df但需要同步，添加该列    
    if '已修改' in value_columns and '已修改' not in campaign_df.columns:
        campaign_df['已修改'] = False
    
    try:
        price_input_indexed = price_input_df.set_index(key_columns)
        
        for idx, row in campaign_df.iterrows():
            # 构建匹配键
            key = tuple(str(row[col]).strip().replace('.0', '') if pd.notnull(row[col]) else '' for col in key_columns)
            
            if key in price_input_indexed.index:
                # 同步值字段
                for col in value_columns:
                    if col in price_input_indexed.columns:
                        campaign_df.at[idx, col] = price_input_indexed.at[key, col]
                
                # 更新价格来源
                if update_price_source and '价格来源' in campaign_df.columns and '已修改' in price_input_indexed.columns:
                    if price_input_indexed.at[key, '已修改']:
                        campaign_df.at[idx, '价格来源'] = '推荐价格'
    except Exception as e:
        st.error(f"同步数据时出错: {str(e)}")
    
    return campaign_df

# 在文件开头添加一个数据验证函数
def validate_required_columns(df, required_columns, df_name="DataFrame"):
    """
    验证DataFrame是否包含所有必要的列
    
    参数:
    df: 要验证的DataFrame
    required_columns: 必需的列名列表
    df_name: DataFrame的名称，用于错误提示
    
    返回:
    (bool, str): (是否有效, 错误信息)
    """
    if df is None:
        return False, f"{df_name}未提供"
    
    missing_columns = [col for col in required_columns if col not in df.columns]
    if missing_columns:
        return False, f"{df_name}缺少必要列: {', '.join(missing_columns)}"
    
    return True, ""

# 修改get_tool_price函数为更高效的向量化版本
def get_tool_price_vectorized(campaign_df, tool_price_df):
    """
    向量化处理SKU价格匹配，替代逐行apply操作
    
    参数:
    campaign_df: 活动价格表DataFrame
    tool_price_df: 工具价格表DataFrame
    
    返回:
    更新后的campaign_df，添加价格和价格来源列
    """
    # 调试信息：输出数据结构
    print(f"活动表包含行数: {len(campaign_df)}")
    print(f"工具价格表包含行数: {len(tool_price_df)}")
    print(f"是否包含SKU列: {SKU_FIELD in campaign_df.columns}")
    print(f"是否包含Parent SKU列: {PARENT_SKU_FIELD in campaign_df.columns}")
    
    # 初始化结果列
    campaign_df[CAMPAIGN_PRICE_FIELD] = np.nan
    campaign_df['价格来源'] = '推荐价格'  # 默认来源为推荐价格
    
    # 创建SKU对应的价格映射字典 - 比逐行查找更高效
    sku_price_dict = dict(zip(
        tool_price_df[TOOL_SKU_FIELD].astype(str).str.strip(),
        tool_price_df[TOOL_PRICE_FIELD]
    ))
    
    # 调试信息：输出字典信息
    print(f"价格字典包含SKU数量: {len(sku_price_dict)}")
    if len(sku_price_dict) > 0:
        # 随机抽样5个
        sample_keys = list(sku_price_dict.keys())[:5]
        print(f"样本SKU: {sample_keys}")
        print(f"样本价格: {[sku_price_dict[k] for k in sample_keys]}")
    
    # 检查nan值
    if 'nan' in sku_price_dict:
        print(f"警告: 价格字典中包含'nan'键，值为: {sku_price_dict['nan']}")
        # 从字典中移除'nan'键，避免错误匹配
        if 'nan' in sku_price_dict:
            del sku_price_dict['nan']
            print("已从价格字典中移除'nan'键")
    
    # 1. 首先尝试直接匹配SKU
    if SKU_FIELD in campaign_df.columns:
        sku_mask = campaign_df[SKU_FIELD].astype(str).str.strip().isin(sku_price_dict.keys())
        if sku_mask.any():
            # 对匹配到的SKU设置价格
            for idx in campaign_df[sku_mask].index:
                sku = str(campaign_df.at[idx, SKU_FIELD]).strip()
                # 排除nan和空字符串
                if sku.lower() == 'nan' or sku == '':
                    continue
                    
                if sku in sku_price_dict:
                    price_val = sku_price_dict[sku]
                    # 修改逻辑：区分有效工具价格和无效工具价格（零或空）
                    if pd.notnull(price_val) and price_val > 0:
                        campaign_df.at[idx, CAMPAIGN_PRICE_FIELD] = price_val
                        campaign_df.at[idx, '价格来源'] = '工具价格'
                    elif pd.notnull(price_val) and price_val == 0:
                        # 价格为零，标记为无效工具价格，仍使用推荐价格
                        campaign_df.at[idx, CAMPAIGN_PRICE_FIELD] = campaign_df.at[idx, CAMPAIGN_RECOMMEND_FIELD]
                        campaign_df.at[idx, '价格来源'] = '无效工具价格(零)'
    
    # 2. 然后尝试匹配Parent SKU (对未匹配到SKU的行)
    if PARENT_SKU_FIELD in campaign_df.columns:
        # 找出还没匹配到价格或标记为无效工具价格的行
        parent_mask = ((campaign_df['价格来源'] == '推荐价格') | 
                       (campaign_df['价格来源'] == '无效工具价格(零)')) & campaign_df[PARENT_SKU_FIELD].notna()
        
        # 添加调试信息
        parent_count = parent_mask.sum()
        print(f"需要尝试Parent SKU匹配的行数: {parent_count}")
        
        if parent_mask.any():
            # 输出一些Parent SKU样本
            parent_sample = campaign_df[parent_mask][PARENT_SKU_FIELD].head(5).tolist()
            print(f"Parent SKU样本: {parent_sample}")
            print(f"这些Parent SKU是否在价格字典中: {[sku in sku_price_dict for sku in parent_sample]}")
            
            # 创建一个字典记录哪些Parent SKU被成功匹配
            parent_matched = {}
            
            # 对匹配到的Parent SKU设置价格
            for idx in campaign_df[parent_mask].index:
                parent_sku = str(campaign_df.at[idx, PARENT_SKU_FIELD]).strip()
                # 排除nan和空字符串
                if parent_sku.lower() == 'nan' or parent_sku == '':
                    continue
                    
                if parent_sku in sku_price_dict:
                    price_val = sku_price_dict[parent_sku]
                    # 修改逻辑：区分有效工具价格和无效工具价格（零或空）
                    if pd.notnull(price_val) and price_val > 0:
                        campaign_df.at[idx, CAMPAIGN_PRICE_FIELD] = price_val
                        campaign_df.at[idx, '价格来源'] = 'Parent工具价格'
                        # 记录匹配成功
                        if parent_sku not in parent_matched:
                            parent_matched[parent_sku] = 1
                        else:
                            parent_matched[parent_sku] += 1
                    elif pd.notnull(price_val) and price_val == 0:
                        # Parent价格为零，也标记为无效工具价格
                        campaign_df.at[idx, CAMPAIGN_PRICE_FIELD] = campaign_df.at[idx, CAMPAIGN_RECOMMEND_FIELD]
                        campaign_df.at[idx, '价格来源'] = '无效Parent工具价格(零)'
            
            # 输出Parent SKU匹配统计
            print(f"通过Parent SKU成功匹配的行数: {sum(parent_matched.values())}")
            print(f"成功匹配的唯一Parent SKU数量: {len(parent_matched)}")
            if len(parent_matched) > 0:
                top_parents = sorted(parent_matched.items(), key=lambda x: x[1], reverse=True)[:5]
                print(f"匹配次数最多的Parent SKU: {top_parents}")
                
                # 检查这些Parent SKU对应的价格
                for parent, _ in top_parents:
                    if parent in sku_price_dict:
                        print(f"Parent SKU {parent} 对应价格: {sku_price_dict[parent]}")
    
    # 3. 最后，对未匹配到的行使用推荐价格
    remaining_mask = (campaign_df['价格来源'] == '推荐价格')
    campaign_df.loc[remaining_mask, CAMPAIGN_PRICE_FIELD] = campaign_df.loc[remaining_mask, CAMPAIGN_RECOMMEND_FIELD]
    
    # 保存是否有需要审查的价格数据
    需要审查的价格条件 = (
        (campaign_df['价格来源'] == '推荐价格') | 
        (campaign_df['价格来源'] == '无效工具价格(零)') | 
        (campaign_df['价格来源'] == '无效Parent工具价格(零)')
    )
    
    return campaign_df

st.set_page_config(page_title="SKU活动价自动匹配与审核工具_v1.0（测试版/开发中）", layout="wide")

st.title("SKU活动价自动匹配与审核工具_v1.2（测试版/开发中）")
st.markdown("""
#### 操作指引：
1. 上传SKU表、工具价格表、活动价格提交表（支持Excel/CSV）。
2. 工具自动匹配并填写活动价格，优先用工具价格表，匹配不到用Parent SKU，再匹配不到用推荐价格。
3. 自动填入推荐价格的行可人工确认或修改。
4. 确认无误后可导出最终表格。
5. **如活动价格提交表第2-3行为备注说明，可在上传时设置跳过的备注行范围（如2-3行为备注，则起始为2，结束为3）。**

> 当前为测试版，功能持续开发中，结果仅供参考。
""")

# 在程序开始处初始化关键变量，避免NameError
campaign_df = None
raw_campaign_df = None
export_df = None
editable_df = None
campaign_file = None
skip_start = 2
skip_end = 3

# 上传SKU表和工具价格表后，均支持选择表头行
sku_df = None
tool_price_df = None
sku_header_row = 3
tool_header_row = 2

col1, col2, col3 = st.columns(3)
with col1:
    sku_file = st.file_uploader("上传SKU表", type=["xlsx", "xls", "csv"], key="sku")
    if sku_file is not None:
        sku_header_row = st.number_input("SKU表表头所在行", min_value=1, max_value=5, value=3, key="sku_header")
        # 先读取为BytesIO，兼容openpyxl
        file_bytes = io.BytesIO(sku_file.read())
        sku_df = strip_columns(pd.read_excel(file_bytes, header=sku_header_row-1))

with col2:
    tool_price_file = st.file_uploader("上传工具价格表", type=["xlsx", "xls", "csv"], key="tool")
    if tool_price_file is not None:
        tool_header_row = st.number_input("工具价格表表头所在行", min_value=1, max_value=5, value=2, key="tool_header")
        tool_price_df = strip_columns(pd.read_excel(tool_price_file, header=tool_header_row-1))

with col3:
    campaign_file = st.file_uploader("上传活动价格提交表", type=["xlsx", "xls", "csv"], key="campaign")
    if campaign_file is not None:
        skip_col1, skip_col2 = st.columns(2)
        with skip_col1:
            skip_start = st.number_input("备注起始行号（从1开始）", min_value=2, max_value=10, value=2, key="skip_start")
        with skip_col2:
            skip_end = st.number_input("备注结束行号（从1开始）", min_value=skip_start, max_value=20, value=3, key="skip_end")
        # 计算需要跳过的行（pandas的skiprows是从0开始的索引）
        skiprows = list(range(skip_start-1, skip_end))
        raw_campaign_df = pd.read_excel(campaign_file, header=0, skiprows=skiprows)  # 原始表格
        campaign_df = strip_columns(raw_campaign_df.copy())  # 用于后续处理
        
        # 调试信息：输出campaign_df的列名
        st.write("### Campaign表列名检查")
        st.write(f"原始列名: {list(campaign_df.columns)}")
        # 检查是否包含必要的列
        required_cols = [CAMPAIGN_PRODUCT_ID, CAMPAIGN_VARIATION_ID, CAMPAIGN_RECOMMEND_FIELD, CAMPAIGN_PRICE_FIELD]
        missing_cols = [col for col in required_cols if col not in campaign_df.columns]
        if missing_cols:
            st.error(f"Campaign表缺少必要列: {', '.join(missing_cols)}")
            st.write("可能的列名映射问题，请检查字段名配置或调整表头")

# 保证用于合并的字段类型一致，并清洗SKU相关字段
for col in [CAMPAIGN_PRODUCT_ID, CAMPAIGN_VARIATION_ID]:
    if campaign_df is not None and col in campaign_df.columns:
        campaign_df[col] = campaign_df[col].astype(str).str.strip()
    if sku_df is not None and col in sku_df.columns:
        sku_df[col] = sku_df[col].astype(str).str.strip()

# 清洗SKU字段，确保内容一致
if sku_df is not None and SKU_FIELD in sku_df.columns:
    sku_df[SKU_FIELD] = sku_df[SKU_FIELD].astype(str).str.strip()
if sku_df is not None and PARENT_SKU_FIELD in sku_df.columns:
    sku_df[PARENT_SKU_FIELD] = sku_df[PARENT_SKU_FIELD].astype(str).str.strip()
if tool_price_df is not None and TOOL_SKU_FIELD in tool_price_df.columns:
    tool_price_df[TOOL_SKU_FIELD] = tool_price_df[TOOL_SKU_FIELD].astype(str).str.strip()

# 新增：ID字段清洗函数，去除小数点（如.0），保证编号匹配
# SKU表
if sku_df is not None:
    for col in [SKU_FIELD, PARENT_SKU_FIELD, CAMPAIGN_PRODUCT_ID, CAMPAIGN_VARIATION_ID]:
        sku_df = clean_id_column(sku_df, col)
# 工具价格表
if tool_price_df is not None:
    tool_price_df = clean_id_column(tool_price_df, TOOL_SKU_FIELD)
# 活动价格表
if campaign_df is not None:
    for col in [CAMPAIGN_PRODUCT_ID, CAMPAIGN_VARIATION_ID]:
        campaign_df = clean_id_column(campaign_df, col)

st.markdown("---")#分隔符

st.subheader('价格确认与导出')

st.markdown('**价格浮动范围设置**（推荐价格的±百分比，默认50%，可自定义）')
price_range_percent = st.number_input('允许价格浮动范围（%）', min_value=0, max_value=100, value=50, step=1)

if sku_df is not None and tool_price_df is not None and campaign_df is not None:
    # 数据验证 - 检查必要字段
    sku_required = [CAMPAIGN_PRODUCT_ID, CAMPAIGN_VARIATION_ID, SKU_FIELD, PARENT_SKU_FIELD]
    tool_required = [TOOL_SKU_FIELD, TOOL_PRICE_FIELD]
    campaign_required = [CAMPAIGN_PRODUCT_ID, CAMPAIGN_VARIATION_ID, CAMPAIGN_RECOMMEND_FIELD]
    
    # 验证各表必要字段
    is_valid_sku, sku_error = validate_required_columns(sku_df, sku_required, "SKU表")
    is_valid_tool, tool_error = validate_required_columns(tool_price_df, tool_required, "工具价格表")
    is_valid_campaign, campaign_error = validate_required_columns(campaign_df, campaign_required, "活动价格提交表")
    
    if not (is_valid_sku and is_valid_tool and is_valid_campaign):
        missing_fields = []
        if not is_valid_sku: missing_fields.append(sku_error)
        if not is_valid_tool: missing_fields.append(tool_error)
        if not is_valid_campaign: missing_fields.append(campaign_error)
        st.error("数据验证失败：\n" + "\n".join(missing_fields))
    else:
        # 合并SKU信息到活动价格表
        sku_id_columns = [CAMPAIGN_PRODUCT_ID, CAMPAIGN_VARIATION_ID]
        sku_merge_columns = [CAMPAIGN_PRODUCT_ID, CAMPAIGN_VARIATION_ID, SKU_FIELD, PARENT_SKU_FIELD]
        campaign_df = campaign_df.merge(sku_df[sku_merge_columns], on=sku_id_columns, how="left")
        
        # 使用向量化方法进行价格匹配
        campaign_df = get_tool_price_vectorized(campaign_df, tool_price_df)
    
    campaign_df['需用户确认'] = campaign_df['价格来源'] == '推荐价格'
    # 调试信息：输出价格来源统计 
    st.write("### 调试信息")
    st.write("价格来源统计:", campaign_df['价格来源'].value_counts().to_dict())
    st.success("自动匹配完成，橙色高亮行为需人工确认/修改：")

    # 可编辑表格过滤：显示推荐价格、无效工具价格、匹配失败的行
    campaign_df['初始推荐价格'] = campaign_df[CAMPAIGN_RECOMMEND_FIELD]
        
    # 保存是否有需要审查的价格数据
    需要审查的价格条件 = (
        (campaign_df['价格来源'] == '推荐价格') | 
        (campaign_df['价格来源'] == '无效工具价格(零)') | 
        (campaign_df['价格来源'] == '无效Parent工具价格(零)')
    )
    has_prices_to_review = 需要审查的价格条件.any()
        
    # 调试信息：输出审查条件统计
    st.write(f"需要审查的价格数量: {需要审查的价格条件.sum()}")
        
    # 筛选条件：包含所有非有效工具价格的数据
    需确认条件 = 需要审查的价格条件  # 推荐价格或无效工具价格
    价格缺失条件 = (campaign_df[CAMPAIGN_PRICE_FIELD].isnull() | (campaign_df[CAMPAIGN_PRICE_FIELD] == ""))
    st.write(f"价格缺失数量: {价格缺失条件.sum()}")
        
    编辑表筛选条件 = 需确认条件 | 价格缺失条件
    st.write(f"编辑表筛选条件匹配数量: {编辑表筛选条件.sum()}")
        
    editable_df = campaign_df[编辑表筛选条件].copy()
    st.write(f"可编辑表格行数: {len(editable_df)}")
        
    # 确保editable_df不为空才执行填充操作
    if not editable_df.empty:
        # 自动填入推荐价格（缺失时）
        价格缺失掩码 = editable_df[CAMPAIGN_PRICE_FIELD].isnull() | (editable_df[CAMPAIGN_PRICE_FIELD] == "")
        if 价格缺失掩码.any():
            editable_df.loc[价格缺失掩码, CAMPAIGN_PRICE_FIELD] = editable_df.loc[价格缺失掩码, CAMPAIGN_RECOMMEND_FIELD].values

    # 新增：标记修改列，仅做视觉标记
    editable_df['标记修改'] = False

    # 列顺序（去掉"标记修改"和"需用户确认"）
    显示列优先顺序 = [
        CAMPAIGN_PRODUCT_ID, CAMPAIGN_VARIATION_ID, '价格来源', CAMPAIGN_RECOMMEND_FIELD,
        CAMPAIGN_PRICE_FIELD, '已修改'
    ]
    editable_cols_order = [col for col in 显示列优先顺序 if col in editable_df.columns] + \
                          [col for col in editable_df.columns if col not in 显示列优先顺序 and col not in ['标记修改', '需用户确认']]

    st.markdown("#### 活动价格审核表（可编辑，仅显示推荐价格/匹配失败）")
        
    # 初始化price_input变量，避免在editable_df为空时未定义
    price_input = pd.DataFrame()
        
    # 修改判断逻辑：只有在完全没有需要审查的价格数据时才显示提示信息
    if not has_prices_to_review and 价格缺失条件.sum() == 0:
        st.info("没有需要人工确认或修改的价格，所有价格已自动匹配完成！")
    else:
        # 添加已人工确认列
        editable_df['已人工确认'] = False
        editable_cols_order = [col for col in 显示列优先顺序 if col in editable_df.columns] + \
                             [col for col in editable_df.columns if col not in 显示列优先顺序 and col not in ['标记修改', '需用户确认']]
        
        # 确保'已人工确认'列在显示列中
        if '已人工确认' not in editable_cols_order and '已人工确认' in editable_df.columns:
            editable_cols_order.append('已人工确认')
        
        # 如果editable_df为空（虽然有推荐价格来源但可能被过滤掉），添加提示
        if editable_df.empty:
            st.warning("筛选后没有数据显示，请检查筛选条件")
        else:
            price_input = st.data_editor(
                editable_df[editable_cols_order],
                use_container_width=True,
                num_rows="dynamic",
                disabled=[col for col in editable_df.columns if col != CAMPAIGN_PRICE_FIELD and col != '已人工确认'],
                hide_index=True,
                key="editable_confirm"
            )
            
            # 检查price_input是否为空
            st.write(f"数据编辑器返回的price_input行数: {len(price_input)}")

    # 检查价格是否被修改
    def check_modified(row):
        try:
            # 确保'初始推荐价格'列存在
            if '初始推荐价格' not in row:
                return False
            return float(row[CAMPAIGN_PRICE_FIELD]) != float(row['初始推荐价格'])
        except (ValueError, TypeError, KeyError):
            # 明确异常类型，避免掩盖其他异常
            try:
                if '初始推荐价格' not in row:
                    return False
                return str(row[CAMPAIGN_PRICE_FIELD]) != str(row['初始推荐价格'])
            except KeyError:
                # 如果发生KeyError，可能是'初始推荐价格'或CAMPAIGN_PRICE_FIELD不存在
                st.error(f"检查修改时发生键错误，行内容: {row}")
                return False
    
    try:
        price_input['已修改'] = price_input.apply(lambda row: check_modified(row), axis=1)
        
        # 确保'已人工确认'列存在于同步数据中
        if '已人工确认' not in price_input.columns:
            price_input['已人工确认'] = False
    except Exception as e:
        st.error(f"处理price_input时出错: {str(e)}")
        # 输出price_input的列信息
        st.write(f"price_input列: {list(price_input.columns)}")

    # 价格有效性校验
    def is_price_valid(row, percent):
        try:
            rec = float(row[CAMPAIGN_RECOMMEND_FIELD])
            price = float(row[CAMPAIGN_PRICE_FIELD])
            min_p = rec * (1 - percent/100)
            max_p = rec * (1 + percent/100)
            return min_p <= price <= max_p
        except (ValueError, TypeError, ZeroDivisionError):
            return False

    # 只有在price_input非空时才执行价格验证
    if not price_input.empty and CAMPAIGN_PRICE_FIELD in price_input.columns:
        price_input['价格有效'] = price_input.apply(lambda row: is_price_valid(row, price_range_percent), axis=1)
        # 使用新增的同步函数替代重复代码
        campaign_df = sync_price_data(
            campaign_df, 
            price_input, 
            key_columns=[CAMPAIGN_PRODUCT_ID, CAMPAIGN_VARIATION_ID],
            value_columns=[CAMPAIGN_PRICE_FIELD, '已修改', '价格有效', '已人工确认'],
            update_price_source=True
        )
        
        # 红色警告提示
        invalid_rows = price_input[~price_input['价格有效']] if '价格有效' in price_input.columns else pd.DataFrame()
        if not invalid_rows.empty:
            st.error(f"有{len(invalid_rows)}行价格超出允许浮动范围，请注意核查！")
    else:
        if '价格有效' not in campaign_df.columns:
            campaign_df['价格有效'] = True

# ----------- 只读高亮表应显示所有匹配结果 -----------
# 确保campaign_df不为None再操作
if campaign_df is not None:
    show_cols = [col for col in campaign_df.columns if col not in ['需用户确认', '初始推荐价格', '已人工确认']]
    
    # 确保数据处理中'已人工确认'列存在，虽然不显示
    if '已人工确认' not in campaign_df.columns:
        campaign_df['已人工确认'] = False
    
    # === 千分位格式化显示 Campaign Price 字段 ===
    show_df = campaign_df[show_cols].copy()
    if CAMPAIGN_PRICE_FIELD in show_df.columns:
        show_df[CAMPAIGN_PRICE_FIELD] = show_df[CAMPAIGN_PRICE_FIELD].apply(
            lambda x: '{:,}'.format(int(x)) if pd.notnull(x) and str(x).strip() != "" else x
        )
    
    # 新增：填充空值，防止AgGrid渲染异常
    show_df = show_df.fillna("")
    
    # 将表格标题从"活动价格审核表（只读高亮，无复选框）"改为"活动价预览表"
    st.markdown("#### 活动价预览表（按价格来源高亮显示）")
    
    # 添加简短说明，帮助用户理解不同颜色的含义
    color_info = """
    <small>
    <span style="color:#1E90FF">■</span> 工具价格 | 
    <span style="color:#20B2AA">■</span> Parent工具价格 | 
    <span style="color:#FFD700">■</span> 推荐价格 | 
    <span style="color:#FF8C00">■</span> 无效工具价格(零) | 
    <span style="color:#FF0000">■</span> 价格缺失(含其他严重错误)
    </small>
    """
    st.markdown(color_info, unsafe_allow_html=True)
        
    # ========== st-aggrid 只读预览表 ==========
    gb = GridOptionsBuilder.from_dataframe(show_df)
    gb.configure_default_column(filter=True, sortable=True)
    gb.configure_grid_options(domLayout='normal')
    # 多色高亮和价格缺失高亮
    cellstyle_jscode = JsCode("""
    function(params) {
        // 价格来源多色高亮
        if (params.colDef.field === '价格来源') {
            if (params.value === '工具价格') {
                return { 'color': 'white', 'backgroundColor': '#1E90FF' }
            }
            if (params.value === 'Parent工具价格') {
                return { 'color': 'white', 'backgroundColor': '#20B2AA' }
            }
            if (params.value === '推荐价格') {
                return { 'color': 'black', 'backgroundColor': '#FFD700' }
            }
            if (params.value === '无效工具价格(零)' || params.value === '无效Parent工具价格(零)') {
                return { 'color': 'white', 'backgroundColor': '#FF8C00' }
            }
        }
        // 活动价格缺失/错误高亮
        if (params.colDef.field === '活动价格' || params.colDef.field === 'Campaign Price') {
            if (params.value === '' || params.value === null || params.value === 0 || params.value === '0' || params.value === 'None') {
                return { 'color': 'white', 'backgroundColor': '#FF0000' }
            }
        }
        return {};
    }
    """)
    if '价格来源' in show_df.columns:
        gb.configure_column('价格来源', cellStyle=cellstyle_jscode)
    if '活动价格' in show_df.columns:
        gb.configure_column('活动价格', cellStyle=cellstyle_jscode)
    elif 'Campaign Price' in show_df.columns:
        gb.configure_column('Campaign Price', cellStyle=cellstyle_jscode)
    gridOptions = gb.build()
    # 汉化菜单
    locale_cn = {
        "page": "页",
        "more": "更多",
        "to": "到",
        "of": "共",
        "next": "下一页",
        "last": "最后一页",
        "first": "第一页",
        "previous": "上一页",
        "loadingOoo": "加载中...",
        "selectAll": "全选",
        "searchOoo": "搜索...",
        "blank": "空",
        "filterOoo": "自定义筛选...",
        "applyFilter": "应用",
        "equals": "等于",
        "notEqual": "不等于",
        "lessThan": "小于",
        "greaterThan": "大于",
        "lessThanOrEqual": "小于等于",
        "greaterThanOrEqual": "大于等于",
        "inRange": "区间",
        "contains": "包含",
        "notContains": "不包含",
        "startsWith": "开头是",
        "endsWith": "结尾是",
        "andCondition": "并且",
        "orCondition": "或者",
        "noRowsToShow": "无数据显示",
        "copy": "复制",
        "copyWithHeaders": "带表头复制",
        "paste": "粘贴",
        "export": "导出",
        "exportToCsv": "导出为CSV",
        "exportToExcel": "导出为Excel",
        "pinColumn": "固定列",
        "valueAggregation": "聚合",
        "autosizeThiscolumn": "自动调整本列宽度",
        "autosizeAllColumns": "自动调整所有列宽度",
        "resetColumns": "重置列",
        "groupBy": "按此列分组",
        "ungroupBy": "取消分组",
        "resetGroup": "重置分组",
        "rowGroupColumnsEmptyMessage": "拖动列到此处进行分组",
        "valueColumnsEmptyMessage": "拖动列到此处进行聚合",
        "pivotMode": "透视模式",
        "groups": "分组",
        "values": "值",
        "pivots": "透视",
        "group": "分组",
        "columnsPanel": "列面板",
        "filters": "筛选",
        "rowGroup": "行分组",
        "rowGroupPanel": "行分组面板",
        "pivot": "透视",
        "pivotPanel": "透视面板",
        "notBlank": "非空",
        "resetFilter": "重置",
        "clearFilter": "清除",
        "cancelFilter": "取消",
        "apply": "应用",
        "cancel": "取消",
        "clear": "清除",
        "textFilter": "文本筛选",
        "numberFilter": "数字筛选",
        "dateFilter": "日期筛选",
        "setFilter": "集合筛选",
        "columns": "列",
        "menu": "菜单",
        "filter": "筛选",
        "deleteCondition": "删除条件",
        "addCondition": "添加条件",
        "filterConditions": "筛选条件",
        "filterValue": "筛选值",
        "filterField": "筛选字段",
        "selectAllSearchResults": "全选搜索结果",
        "search": "搜索",
        "noMatches": "无匹配项",
        "toolPanelColumns": "列",
        "toolPanelFilters": "筛选",
    }
    AgGrid(
        show_df,
        gridOptions=gridOptions,
        enable_enterprise_modules=False,
        fit_columns_on_grid_load=True,
        allow_unsafe_jscode=True,
        theme='streamlit',
        update_mode='NO_UPDATE',
        localeText=locale_cn
    )

# === 更高效的价格设置方法 ===
def apply_campaign_price_to_export(export_df, campaign_df, key_columns):
    """
    更高效地将活动价格应用到导出DataFrame
    
    参数:
    export_df: 导出用的DataFrame
    campaign_df: 包含价格和来源信息的DataFrame
    key_columns: 用于匹配两个DataFrame的键列
    
    返回:
    更新后的export_df
    """
    # 准备需要的字段
    if len(key_columns) == 0:
        st.warning("没有找到合适的键列进行数据匹配")
        return export_df
    
    # 确保campaign_df中包含必要的列
    required_columns = [CAMPAIGN_PRICE_FIELD, '价格来源']
    for col in required_columns:
        if col not in campaign_df.columns:
            st.error(f"数据处理错误: campaign_df中缺少必要列 '{col}'")
            return export_df
    
    # 确保必要的列存在，如不存在则创建
    if '已修改' not in campaign_df.columns:
        campaign_df['已修改'] = False
    
    if '已人工确认' not in campaign_df.columns:
        campaign_df['已人工确认'] = False
    
    # 创建匹配用的键
    campaign_df['匹配键'] = campaign_df[key_columns].astype(str).apply(
        lambda x: '-'.join([str(i).strip().replace('.0', '') for i in x]), axis=1
    )
    export_df['匹配键'] = export_df[key_columns].astype(str).apply(
        lambda x: '-'.join([str(i).strip().replace('.0', '') for i in x]), axis=1
    )
    
    # 提取要复制的字段，增加已人工确认列
    campaign_slim = campaign_df[['匹配键', CAMPAIGN_PRICE_FIELD, '价格来源', '已修改', '已人工确认']].copy()
    
    # 使用merge代替循环 - 更高效
    result_df = export_df.merge(campaign_slim, on='匹配键', how='left', suffixes=('', '_new'))
    
    # 更新价格
    if CAMPAIGN_PRICE_FIELD + '_new' in result_df.columns:
        result_df[CAMPAIGN_PRICE_FIELD] = result_df[CAMPAIGN_PRICE_FIELD + '_new'].fillna(result_df[CAMPAIGN_PRICE_FIELD])
    
    # 设置价格标记
    # 先检查合并后的列是否存在
    if '价格来源' not in result_df.columns:
        result_df['价格标记'] = ''
        st.warning("价格来源信息缺失，无法设置详细价格标记")
    else:
        # 定义条件，使用.fillna确保没有NaN值
        条件_工具价格 = (result_df['价格来源'] == '工具价格').fillna(False)
        条件_Parent价格 = (result_df['价格来源'] == 'Parent工具价格').fillna(False) 
        条件_推荐价格 = (result_df['价格来源'] == '推荐价格').fillna(False)
        条件_无效工具价格 = (result_df['价格来源'] == '无效工具价格(零)').fillna(False)
        条件_无效Parent工具价格 = (result_df['价格来源'] == '无效Parent工具价格(零)').fillna(False)
        
        # 检查'已修改'和'已人工确认'列是否存在
        条件_已修改 = result_df['已修改'].fillna(False) if '已修改' in result_df.columns else pd.Series(False, index=result_df.index)
        条件_已人工确认 = result_df['已人工确认'].fillna(False) if '已人工确认' in result_df.columns else pd.Series(False, index=result_df.index)
        
        # 创建标记列，增加人工确认信息
        result_df['价格标记'] = ''
        result_df.loc[条件_工具价格, '价格标记'] = '工具价格'
        result_df.loc[条件_Parent价格, '价格标记'] = 'Parent工具价格'
        result_df.loc[条件_推荐价格 & ~条件_已修改 & ~条件_已人工确认, '价格标记'] = '推荐价格'
        result_df.loc[条件_推荐价格 & 条件_已修改, '价格标记'] = '推荐价格（已手动更改）'
        result_df.loc[条件_推荐价格 & ~条件_已修改 & 条件_已人工确认, '价格标记'] = '推荐价格（已人工确认）'
        result_df.loc[条件_推荐价格 & 条件_已修改 & 条件_已人工确认, '价格标记'] = '推荐价格（已手动更改并确认）'
        # 无效工具价格的标记
        result_df.loc[条件_无效工具价格 & ~条件_已修改 & ~条件_已人工确认, '价格标记'] = '无效工具价格(零)'
        result_df.loc[条件_无效工具价格 & 条件_已修改, '价格标记'] = '无效工具价格(零)（已手动更改）'
        result_df.loc[条件_无效工具价格 & ~条件_已修改 & 条件_已人工确认, '价格标记'] = '无效工具价格(零)（已人工确认）'
        result_df.loc[条件_无效工具价格 & 条件_已修改 & 条件_已人工确认, '价格标记'] = '无效工具价格(零)（已手动更改并确认）'
        # 无效Parent工具价格的标记
        result_df.loc[条件_无效Parent工具价格 & ~条件_已修改 & ~条件_已人工确认, '价格标记'] = '无效Parent工具价格(零)'
        result_df.loc[条件_无效Parent工具价格 & 条件_已修改, '价格标记'] = '无效Parent工具价格(零)（已手动更改）'
        result_df.loc[条件_无效Parent工具价格 & ~条件_已修改 & 条件_已人工确认, '价格标记'] = '无效Parent工具价格(零)（已人工确认）'
        result_df.loc[条件_无效Parent工具价格 & 条件_已修改 & 条件_已人工确认, '价格标记'] = '无效Parent工具价格(零)（已手动更改并确认）'
        
        # 价格缺失或严重错误情况
        价格缺失条件 = (result_df[CAMPAIGN_PRICE_FIELD].isnull() | 
                    (result_df[CAMPAIGN_PRICE_FIELD] == "") | 
                    (result_df[CAMPAIGN_PRICE_FIELD] == 0))
        if 价格缺失条件.any():
            result_df.loc[价格缺失条件, '价格标记'] = '价格缺失(含其他严重错误)'
    
    # 删除临时列和不需要的merge结果列
    drop_cols = ['匹配键']
    for col in [CAMPAIGN_PRICE_FIELD + '_new', '价格来源', '已修改', '已人工确认']:
        if col in result_df.columns:
            drop_cols.append(col)
    
    result_df = result_df.drop(columns=drop_cols)
    
    return result_df

# 新增：导出时只写价格，并在末尾添加标记信息
if raw_campaign_df is not None:
    export_df = raw_campaign_df.copy()
    
    # 唯一键
    sku_id_列 = []
    if CAMPAIGN_PRODUCT_ID in export_df.columns:
        sku_id_列.append(CAMPAIGN_PRODUCT_ID)
    if CAMPAIGN_VARIATION_ID in export_df.columns:
        sku_id_列.append(CAMPAIGN_VARIATION_ID)
    
    # 检查导出数据的有效性
    if not sku_id_列:
        st.error(f"导出表缺少必要的ID列 {CAMPAIGN_PRODUCT_ID} 或 {CAMPAIGN_VARIATION_ID}")
    elif campaign_df is not None:
        # 使用更高效的方法更新价格和标记
        export_df = apply_campaign_price_to_export(export_df, campaign_df, sku_id_列)
else:
    export_df = None
    st.warning("未加载活动价格提交表，无法导出数据")

# 确保editable_df已定义，避免NameError
if 'editable_df' not in locals():
    editable_df = None

# === 在此处格式化价格字段为整数 ===
for df in [campaign_df, editable_df, export_df]:
    if df is not None and CAMPAIGN_PRICE_FIELD in df.columns:
        try:
            df[CAMPAIGN_PRICE_FIELD] = df[CAMPAIGN_PRICE_FIELD].apply(
                lambda x: int(float(x)) if pd.notnull(x) and str(x).strip() != "" else x
            )
        except (ValueError, TypeError):
            # 捕获可能出现的类型转换错误
            st.warning(f"价格字段包含无法转换为整数的值，请检查数据")
    if df is not None and CAMPAIGN_RECOMMEND_FIELD in df.columns:
        try:
            df[CAMPAIGN_RECOMMEND_FIELD] = df[CAMPAIGN_RECOMMEND_FIELD].apply(
                lambda x: int(float(x)) if pd.notnull(x) and str(x).strip() != "" else x
            )
        except (ValueError, TypeError):
            st.warning(f"推荐价格字段包含无法转换为整数的值，请检查数据")

# 拼接remark行
if 'skip_end' not in locals() or skip_end is None:
    skip_end = 0
    
remark_rows = skip_end
try:
    if campaign_file is not None and export_df is not None:
        remark_df = pd.read_excel(campaign_file, header=None, nrows=remark_rows)
        # remark_df只赋值它实际有的列名
        remark_col_num = remark_df.shape[1]
        remark_df.columns = list(export_df.columns)[:remark_col_num]
        # 补齐缺失的列
        for col in export_df.columns[remark_col_num:]:
            remark_df[col] = ""
        # 对齐顺序
        remark_df = remark_df[export_df.columns]
        
        final_df = pd.concat([remark_df, export_df], ignore_index=True)
    else:
        final_df = export_df if export_df is not None else pd.DataFrame()
except Exception as e:
    st.warning(f"处理备注行时出错: {str(e)}，已忽略备注行")
    final_df = export_df.copy() if export_df is not None else pd.DataFrame()

# === 只保留一处导出按钮和逻辑 ===
col_head, col_mark = st.columns(2)
with col_head:
    header_row_input = st.number_input(
        "活动价格提交表表头实际所在行号（从1开始）",
        min_value=1,
        max_value=50,
        value=1,
        key="campaign_header_row"
    )
with col_mark:
    price_mark_col = st.number_input(
        "价格标记插入列号（默认16，强制写入该列，原有内容会被覆盖）",
        min_value=1,
        max_value=50,
        value=16,
        key="price_mark_col"
    )
header_row = header_row_input  # 用户视角，直接用输入值，不做-1

# 用 session_state 缓存导出内容
if 'export_output' not in st.session_state:
    st.session_state['export_output'] = None

# 确保campaign_file已定义
if 'campaign_file' not in locals() or campaign_file is None:
    campaign_file = None

if st.button("生成最终活动价格表（Excel）"):
    if campaign_file is None:
        st.error("请先上传活动价格提交表")
    elif export_df is None:
        st.error("没有可导出的数据")
    else:
        try:
            wb = openpyxl.load_workbook(campaign_file)
            ws = wb.active

            # 读取表头行（openpyxl行号从1开始）
            col_names = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[header_row]]

            if CAMPAIGN_PRICE_FIELD not in col_names:
                st.error(f"列名 '{CAMPAIGN_PRICE_FIELD}' 不在表头中，请检查表头行或列名是否正确！")
            else:
                price_col_idx = col_names.index(CAMPAIGN_PRICE_FIELD) + 1
                # 用用户选择的列号插入"价格标记"表头
                ws.cell(row=header_row, column=price_mark_col, value="价格标记")

                # 修复: 正确计算数据写入的起始行
                # 1. 如果备注行在表头之前，数据起始行 = 表头行 + 1
                # 2. 如果备注行在表头之后，数据起始行 = 表头行 + 1 + (备注结束行 - 表头行)
                data_start_row = header_row + 1
                if skip_end > header_row:
                    data_start_row += (skip_end - header_row)
                
                # 写入数据（按正确的行号计算）
                for idx, row in export_df.iterrows():
                    excel_row = data_start_row + idx
                    if CAMPAIGN_PRICE_FIELD in row and '价格标记' in row:
                        ws.cell(row=excel_row, column=price_col_idx, value=row[CAMPAIGN_PRICE_FIELD])
                        ws.cell(row=excel_row, column=price_mark_col, value=row['价格标记'])

                # 保存到内存并缓存
                with BytesIO() as output:
                    wb.save(output)
                    st.session_state['export_output'] = output.getvalue()
                    
                st.success("已成功生成Excel文件，请点击下方按钮下载")
        except Exception as e:
            st.error(f"生成Excel文件时出错: {str(e)}")

# 只显示一个下载按钮
if st.session_state.get('export_output'):
    st.download_button(
        label="下载最终活动价格表（Excel）",
        data=st.session_state['export_output'],
        file_name="最终活动价格表.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
else:
    st.warning("请上传SKU表、工具价格表和活动价格提交表，三表齐全后自动处理！")

st.markdown("---")

# 添加 main 函数，作为程序入口点
def main():
    # Streamlit 已经自动运行了应用程序，所以这里不需要额外操作
    pass

if __name__ == "__main__":
    main()